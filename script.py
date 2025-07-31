from urllib.parse import urlparse
from datetime import date, timedelta
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import RunReportRequest, Filter, FilterExpression, Dimension, Metric
from google.oauth2 import service_account
import pandas as pd
import openpyxl
import re
import requests
from openpyxl.styles import Alignment

import sys
import os

def resource_path(rel_path):
    if getattr(sys, 'frozen', False):
        # bundle the folder together
        base_path = sys._MEIPASS
    else:
        # running in normal Python
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, rel_path)

def normalize_path(path):
    path = re.sub(r'//+', '/', path)
    path = re.sub(r'(\/)?index\.html$', '', path, flags=re.IGNORECASE)
    if not path.endswith('/'):
        path += '/'
    return path

def clean_page_title(title):
    if ' - ' in title:
        return title.split(' - ', 1)[0].strip()
    return title.strip()

def get_user_input():
    """Get user input for URLs and file naming preferences"""
    print("=== PAGE INVENTORY TOOL ===\n")
    
    # Get URLs
    urls = []
    print("Enter department URLs (one per line, press Enter twice when done):")
    while True:
        url = input("URL (or press Enter to finish): ").strip()
        if not url:
            break
        urls.append(url)
    
    if not urls:
        print("No URLs provided. Exiting.")
        return None, None
    
    # Get file naming preference
    print("\nFile naming options:")
    print("1. Use department name from URL (e.g., 'biophysics_analytics.xlsx')")
    print("2. Use custom prefix (e.g., 'my_prefix_biophysics.xlsx')")
    print("3. Use custom name for each file")
    
    choice = input("Choose option (1-3): ").strip()
    
    if choice == "1":
        return urls, "auto"
    elif choice == "2":
        prefix = input("Enter prefix: ").strip()
        return urls, f"prefix_{prefix}"
    else:
        return urls, "custom"

def generate_filename(url, naming_mode, custom_names=None):
    """Generate filename based on URL and naming preference"""
    parsed_url = urlparse(url)
    dept_name = parsed_url.path.strip('/').replace('/', '_')
    
    if naming_mode == "auto":
        return f"{dept_name}_analytics.xlsx"
    elif naming_mode.startswith("prefix_"):
        prefix = naming_mode.replace("prefix_", "")
        return f"{prefix}_{dept_name}.xlsx"
    elif naming_mode == "custom" and custom_names:
        return custom_names.get(url, f"{dept_name}_analytics.xlsx")
    else:
        return f"{dept_name}_analytics.xlsx"

def fetch_analytics_data(client, dept_path, start_date, end_date, property_id):
    """Fetch analytics data for a department"""
    request = RunReportRequest(
        property="properties/" + property_id,
        dimensions=[
            Dimension(name="pagePath"),
            Dimension(name="pageTitle")
        ],
        metrics=[
            Metric(name="screenPageViews"),
            Metric(name="activeUsers"),
            Metric(name="userEngagementDuration"),
            Metric(name="bounceRate"),
            Metric(name="eventCount"),
        ],
        date_ranges=[{"start_date": start_date, "end_date": end_date}],
        dimension_filter=FilterExpression(
            filter=Filter(
                field_name="pagePath",
                string_filter={"value": dept_path, "match_type": "BEGINS_WITH"}
            )
        ),
    )
    
    return client.run_report(request)

def process_analytics_data(resp, base_url):
    """Process raw analytics data into structured format"""
    data = []
    for idx in range(len(resp.rows)):
        row = resp.rows[idx]
        try:
            raw_path = row.dimension_values[0].value
            norm_path = normalize_path(raw_path)
            title = clean_page_title(row.dimension_values[1].value)
            pageviews = int(row.metric_values[0].value) if row.metric_values[0].value != "" else 0
            users = int(row.metric_values[1].value) if row.metric_values[1].value != "" else 0
            engagement_time = float(row.metric_values[2].value) if row.metric_values[2].value not in ["", None] else 0.0
            bounce_rate_raw = float(row.metric_values[3].value) if row.metric_values[3].value not in ["", None] else 0.0
            bounce_rate = bounce_rate_raw * 100
            event_count = int(row.metric_values[4].value) if row.metric_values[4].value != "" else 0
            link = base_url + norm_path

            data.append({
                "Page Title": title,
                "URL": link,
                "Normalized Path": norm_path,
                "Views": pageviews,
                "Users": users,
                "Engagement Time (sec)": round(engagement_time, 2),
                "Bounce Rate (%)": bounce_rate,
                "Event Count": event_count,
            })
        except Exception as e:
            print(f"Error in row processing: {e}")
    
    return data

def analyze_pages(grouped_data):
    """Analyze pages and create top 20 and pages to review lists"""
    # Create top 20 pages
    top_20 = grouped_data.sort_values(by="Views", ascending=False)
    if len(top_20) > 20:
        top_20 = top_20.head(20)
    
    # Identify pages to review
    pages_to_review_data = []
    for index, row in grouped_data.iterrows():
        reasons = []
        actions = []

        # Condition 1: Low Page-views
        if row["Views"] <= 25:
            reasons.append("Low page-views")
            actions.append("Remove the page if it's no longer needed, or • Improve its visibility: add links from high-traffic pages or menus.")

        # Condition 2: High Bounce Rate
        if row["Bounce Rate (%)"] >= 45:
            reasons.append("High bounce rate")
            actions.append("Strengthen the \"first impression\": refine the page title, intro sentence, and hero image so they immediately match user intent.")

        # Condition 3: Long Engagement Time Per View
        if row["Engagement Time Per View"] > 60:
            reasons.append("Avg. engagement > 60 s")
            actions.append("Tighten the page: focus on one topic, trim long sections, add clear sub-heads and bullets for quicker scanning.")

        if reasons:
            page_data = row.to_dict()
            page_data["Reason"] = " | ".join(reasons)
            page_data["Suggested Action"] = " | ".join(actions)
            pages_to_review_data.append(page_data)

    to_remove = pd.DataFrame(pages_to_review_data)
    if not to_remove.empty:
        review_columns = ["Page Title", "URL", "Reason", "Suggested Action", "Views", "Users", "Bounce Rate (%)", "Engagement Time Per View", "Event Count", "Views per User"]
        to_remove = to_remove[review_columns]
    
    return top_20, to_remove

def get_ai_insights(grouped_data, section_traffic_percentage, overall_stats):
    """Get AI-generated insights using Gemini API"""
    instructions = (
        "INSTRUCTIONS:\n"
        "- 'Top 20 Pages' tab: Most visited pages in this department.\n"
        "- 'Pages to Review' tab: Pages with low or poor engagement; consider reviewing for updates, consolidation, or removal.\n"
        "- 'All Pages' tab: Full analytics for every tracked page in this department.\n"
        "- 'Summary' tab: Automated high-level advice for improving your section.\n"
    )

    formatted_summary = (
        f"{instructions}\n"
        f"SECTION TRAFFIC PERCENTAGE:\n"
        f"- This department/section accounts for {section_traffic_percentage}% of all tracked site traffic.\n\n"
        f"SUMMARY STATISTICS:\n"
        f"- Total pages: {overall_stats['total_pages']}\n"
        f"- Total views: {overall_stats['total_views']}\n"
        f"- Average views per page: {overall_stats['average_views']:.2f}\n"
        f"- Average users per page: {overall_stats['average_users']:.2f}\n"
        f"- Average engagement time per view (sec): {overall_stats['average_engagement_time_per_view']:.2f}\n"
        f"- Average bounce rate: {overall_stats['average_bounce_rate']:.2f}%\n"
        f"- Pages with high bounce rate (>80%): {overall_stats['pages_with_high_bounce']}\n"
        f"- Pages with low views (<10): {overall_stats['pages_with_low_views']}\n\n"
        f"Top 5 most viewed pages:\n"
    )
    for page in overall_stats["top_5_pages"]:
        formatted_summary += f"    - {page['Page Title']} ({page['Views']} views)\n"
    formatted_summary += "Bottom 5 least viewed pages:\n"
    for page in overall_stats["bottom_5_pages"]:
        formatted_summary += f"    - {page['Page Title']} ({page['Views']} views)\n"

    # Clean up formatting
    formatted_summary = formatted_summary.replace("**", "")
    formatted_summary = re.sub(r"#+\s*", "", formatted_summary)
    
    prompt = (
        formatted_summary +
        "\nBased on this data, provide a very specific summary of what is working well and what is not, "
        "included any key trends or issues. Make your analysis and recommendations by using BOTH the page titles and all available analytics fields: views, users, bounce rate, event count, views per user, and engagement time per view (in seconds). "
        "For each observation, clearly explain WHY you think a particular type of content or topic is performing well or underperforming, making logical inferences from both the title and the analytics. "
        "The recommendations need to be meaningful and actionable for people with varying degrees of website knowledge and front end development skill. "
        "Do not make row-by-row suggestions. Instead, group your analysis and recommendations by content themes, patterns, or page types (not individual pages). "
        "Structure your response under these headings:\n"
        "WHAT'S WORKING\n"
        "WHAT'S NOT WORKING\n"
        "RECOMMENDATIONS\n"
        "Make your analysis as specific and actionable as possible based only on the titles and analytics provided."
    )

    api_key = 'AIzaSyCJZYaI9TyOPbJvhjmH4mLcJp7h0qNd_zI'
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    payload = {
        "contents": [
            {"parts": [{"text": prompt}]}
        ]
    }

    try:
        response = requests.post(url, json=payload, timeout=60)
        if response.status_code == 200:
            res_json = response.json()
            try:
                gemini_reply = res_json["candidates"][0]["content"]["parts"][0]["text"]
            except Exception as nested_e:
                print("Could not extract text from Gemini API:", nested_e)
                gemini_reply = "Failed to parse Gemini API output."
        else:
            print("Gemini API error:", response.status_code, response.text)
            gemini_reply = "API error: " + str(response.status_code)
    except Exception as e:
        gemini_reply = "Error retrieving Gemini suggestions: " + str(e)

    # Clean up Gemini output
    gemini_reply = gemini_reply.replace("**", "")
    gemini_reply = re.sub(r"#+\s*", "", gemini_reply)
    
    return gemini_reply

def format_excel_file(filename, top_20, to_remove, grouped_data, ai_summary):
    """Create and format the Excel file with all sheets"""
    try:
        # Write data to Excel
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        top_20.to_excel(writer, sheet_name='Top 20 Pages', index=False)
        to_remove.to_excel(writer, sheet_name='Pages to Review', index=False)
        grouped_data.to_excel(writer, sheet_name='All Pages', index=False)
        writer.close()
        
        # Add AI summary
        df_gemini = pd.DataFrame([{"Summary": ai_summary}])
        writer2 = pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace")
        df_gemini.to_excel(writer2, sheet_name="Summary", index=False)
        writer2.close()
        
        # Format the Excel file
        wb = openpyxl.load_workbook(filename)
        
        # Move Summary sheet to first position
        ws_summary = wb["Summary"]
        wb._sheets.remove(ws_summary)
        wb._sheets.insert(0, ws_summary)
        ws = ws_summary
        ws.column_dimensions['A'].width = 120
        if ws.max_row > 1:
            cell = ws['A2']
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-fit all sheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                col = list(col)
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max_length + 2
            # Wrap text for headers
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error formatting Excel file {filename}: {e}")
        return False

def process_single_department(url, client, start_date, end_date, filename, property_id):
    """Process a single department URL and generate its Excel file"""
    print(f"\nProcessing: {url}")
    
    try:
        # Parse URL and get department path
        parsed_url = urlparse(url)
        dept_path = normalize_path(parsed_url.path)
        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
        
        print(f"Department path: {dept_path}")
        
        # Fetch analytics data
        resp = fetch_analytics_data(client, dept_path, start_date, end_date, property_id)
        
        if not resp.rows:
            print(f"No data found for {url}")
            return False
        
        # Process the data
        data = process_analytics_data(resp, base_url)
        
        if not data:
            print(f"No valid data found for {url}")
            return False
        
        # Create DataFrame and group data
        df = pd.DataFrame(data)
        agg_dict = {
            "Page Title": "first",
            "Views": "sum",
            "Users": "sum",
            "Engagement Time (sec)": "sum",
            "Bounce Rate (%)": "mean",
            "Event Count": "sum"
        }
        
        grouped = df.groupby(["Normalized Path", "URL"], as_index=False).agg(agg_dict)
        
        # Calculate derived metrics
        grouped["Views per User"] = grouped.apply(
            lambda row: round(row["Views"] / row["Users"], 2) if row["Users"] != 0 else 0, axis=1
        )
        grouped["Engagement Time Per View"] = grouped.apply(
            lambda row: round(row["Engagement Time (sec)"] / row["Views"], 2) if row["Views"] != 0 else 0, axis=1
        )
        
        # Clean up columns
        if "Engagement Time (sec)" in grouped.columns:
            grouped = grouped.drop(columns=["Engagement Time (sec)"])
        if "Normalized Path" in grouped.columns:
            grouped = grouped.drop(columns=["Normalized Path"])
        
        # Remove error pages
        grouped = grouped[grouped["Page Title"] != "Oops! We can't seem to find that page."]
        
        # Analyze pages
        top_20, to_remove = analyze_pages(grouped)
        
        # Get total site views for percentage calculation
        request_total = RunReportRequest(
            property="properties/" + property_id,
            metrics=[Metric(name="screenPageViews")],
            date_ranges=[{"start_date": start_date, "end_date": end_date}]
        )
        resp_total = client.run_report(request_total)
        total_site_views = int(resp_total.rows[0].metric_values[0].value) if resp_total.rows else 0
        
        section_views = grouped["Views"].sum()
        section_traffic_percentage = round((section_views / total_site_views) * 100, 2) if total_site_views > 0 else 0.0
        
        # Calculate overall statistics
        overall_stats = {
            "total_pages": len(grouped),
            "total_views": grouped["Views"].sum(),
            "average_views": grouped["Views"].mean(),
            "average_users": grouped["Users"].mean(),
            "average_engagement_time_per_view": grouped["Engagement Time Per View"].mean(),
            "average_bounce_rate": grouped["Bounce Rate (%)"].mean(),
            "pages_with_high_bounce": grouped[grouped["Bounce Rate (%)"] > 80].shape[0],
            "pages_with_low_views": grouped[grouped["Views"] < 10].shape[0],
            "top_5_pages": grouped.sort_values(by="Views", ascending=False).head(5)[["Page Title", "Views"]].to_dict('records'),
            "bottom_5_pages": grouped.sort_values(by="Views", ascending=True).head(5)[["Page Title", "Views"]].to_dict('records'),
        }
        
        # Get AI insights
        ai_summary = get_ai_insights(grouped, section_traffic_percentage, overall_stats)
        
        # Create Excel file
        success = format_excel_file(filename, top_20, to_remove, grouped, ai_summary)
        
        if success:
            print(f"✓ Successfully created: {filename}")
            return True
        else:
            print(f"✗ Failed to create: {filename}")
            return False
            
    except Exception as e:
        print(f"✗ Error processing {url}: {e}")
        return False

def main():
    """Main function to run the batch processing"""
    # Configuration
    PROPERTY_ID = "319028439"
    KEY_PATH = resource_path("credentials.json")
    
    # Get user input
    urls_and_naming = get_user_input()
    if not urls_and_naming:
        return
    
    urls, naming_mode = urls_and_naming
    
    # Get custom names if needed
    custom_names = {}
    if naming_mode == "custom":
        print("\nEnter custom names for each file:")
        for url in urls:
            custom_name = input(f"Name for {url}: ").strip()
            if custom_name:
                if not custom_name.endswith('.xlsx'):
                    custom_name += '.xlsx'
                custom_names[url] = custom_name
    
    # Set up Google Analytics client
    try:
        creds = service_account.Credentials.from_service_account_file(KEY_PATH)
        client = BetaAnalyticsDataClient(credentials=creds)
    except Exception as e:
        print(f"Error setting up Google Analytics client: {e}")
        print("Make sure credentials.json is in the same directory as the script.")
        return
    
    # Set date range
    start_date = str(date.today() - timedelta(days=365))
    end_date = "today"
    
    print(f"\nProcessing {len(urls)} departments...")
    print(f"Date range: {start_date} to {end_date}")
    
    # Process each URL
    successful_files = []
    failed_urls = []
    
    for url in urls:
        filename = generate_filename(url, naming_mode, custom_names)
        
        if process_single_department(url, client, start_date, end_date, filename, PROPERTY_ID):
            successful_files.append(filename)
        else:
            failed_urls.append(url)
    
    # Summary
    print(f"\n=== PROCESSING COMPLETE ===")
    print(f"Successfully created: {len(successful_files)} files")
    print(f"Failed: {len(failed_urls)} URLs")
    
    if successful_files:
        print(f"\nCreated files:")
        for filename in successful_files:
            print(f"  - {filename}")
    
    if failed_urls:
        print(f"\nFailed URLs:")
        for url in failed_urls:
            print(f"  - {url}")
    
    input("\nPress Enter to exit.")

if __name__ == "__main__":
    main()